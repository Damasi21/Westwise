import json
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch
from urllib.error import URLError

from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    CadastroOmie,
    CategoriaOmie,
    ContaCorrenteOmie,
    ContaPagarOmie,
    ContaReceberOmie,
    ContaDRE,
    ContratoItemOmie,
    ContratoOmie,
    DepartamentoOmie,
    Empresa,
    EmpresaUsuario,
    IntegracaoOmie,
    LancamentoContaCorrenteOmie,
    MetaVendedorComercial,
    MovimentoFinanceiroOmie,
    OrdemServicoItemOmie,
    OrdemServicoOmie,
    PedidoItemOmie,
    PedidoOmie,
    PosicaoEstoqueOmie,
    ProdutoOmie,
    ProjetoOmie,
    ServicoOmie,
    SincronizacaoOmie,
    TipoContaCorrenteOmie,
    VendedorOmie,
)
from .omie import (
    consultar_clientes,
    consultar_contratos,
    consultar_contas_correntes,
    consultar_contas_pagar,
    consultar_contas_receber,
    consultar_extrato_conta_corrente,
    consultar_lancamentos_conta_corrente,
    consultar_movimentos_financeiros,
    consultar_ordens_servico,
    consultar_pedidos,
    consultar_posicoes_estoque,
    consultar_produtos,
    consultar_resumo_financas,
    consultar_servicos,
    consultar_tipos_conta_corrente,
    consultar_vendedores,
    executar_sincronizacao_omie,
)


def arquivo_xlsx(nome_aba, cabecalho, linhas):
    workbook = Workbook()
    ws = workbook.active
    ws.title = nome_aba
    ws.append(cabecalho)
    for linha in linhas:
        ws.append(linha)
    conteudo = BytesIO()
    workbook.save(conteudo)
    return SimpleUploadedFile(
        "planilha.xlsx",
        conteudo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


class ListaEmpresasTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="cliente", password="senha-segura"
        )
        self.empresa_permitida = Empresa.objects.create(
            nome="Empresa Permitida Ltda",
            nome_fantasia="Empresa Permitida",
            cnpj="00.000.000/0001-01",
        )
        self.empresa_bloqueada = Empresa.objects.create(
            nome="Empresa Bloqueada Ltda",
            nome_fantasia="Empresa Bloqueada",
            cnpj="00.000.000/0001-02",
        )
        EmpresaUsuario.objects.create(
            empresa=self.empresa_permitida,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.VISUALIZADOR,
        )

    def test_usuario_anonimo_e_redirecionado_para_login(self):
        response = self.client.get(reverse("empresas:lista"))
        self.assertRedirects(
            response,
            f"{reverse('accounts:login')}?next={reverse('empresas:lista')}",
        )

    def test_cliente_visualiza_apenas_empresa_vinculada(self):
        self.client.force_login(self.usuario)
        response = self.client.get(reverse("empresas:lista"))

        self.assertContains(response, self.empresa_permitida.nome_fantasia)
        self.assertNotContains(response, self.empresa_bloqueada.nome_fantasia)
        self.assertNotContains(response, "Configurações")
        self.assertContains(response, "Insight Wise")
        self.assertContains(response, "/media/insight_wise_login.jpeg")


class UsuariosEmpresaTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin_cliente = User.objects.create_user(
            username="admin_cliente",
            password="senha-segura",
        )
        self.gerente = User.objects.create_user(
            username="gerente_cliente",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa Usuarios Ltda",
            nome_fantasia="Empresa Usuarios",
            cnpj="00.000.000/0001-99",
        )
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.admin_cliente,
            papel=EmpresaUsuario.Papel.ADMINISTRADOR,
        )
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.gerente,
            papel=EmpresaUsuario.Papel.GESTOR,
            areas_permitidas=["financeiro"],
        )

    def test_card_usuarios_aparece_em_parametros(self):
        self.client.force_login(self.admin_cliente)

        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertContains(response, "Usuários")
        self.assertContains(
            response,
            reverse("dashboards:usuarios", kwargs={"empresa_slug": self.empresa.slug}),
        )

    def test_administrador_cadastra_usuario_com_permissoes(self):
        self.client.force_login(self.admin_cliente)

        response = self.client.post(
            reverse("dashboards:usuarios", kwargs={"empresa_slug": self.empresa.slug}),
            {
                "username": "analista_financeiro",
                "first_name": "Ana",
                "email": "ana@empresa.com.br",
                "password": "senha-segura",
                "papel": EmpresaUsuario.Papel.VISUALIZADOR,
                "areas_permitidas": ["financeiro"],
                "dashboards_permitidos": ["financeiro:visao-geral"],
                "ativo": "on",
            },
        )

        self.assertRedirects(
            response,
            reverse("dashboards:usuarios", kwargs={"empresa_slug": self.empresa.slug}),
        )
        vinculo = EmpresaUsuario.objects.get(usuario__username="analista_financeiro")
        self.assertEqual(vinculo.empresa, self.empresa)
        self.assertEqual(vinculo.papel, EmpresaUsuario.Papel.VISUALIZADOR)
        self.assertEqual(vinculo.areas_permitidas, ["financeiro"])
        self.assertEqual(vinculo.dashboards_permitidos, ["financeiro:visao-geral"])

    def test_gerente_nao_edita_administrador(self):
        vinculo_admin = EmpresaUsuario.objects.get(usuario=self.admin_cliente)
        self.client.force_login(self.gerente)

        response = self.client.get(
            reverse("dashboards:usuarios", kwargs={"empresa_slug": self.empresa.slug}),
            {"editar": vinculo_admin.pk},
        )

        self.assertEqual(response.status_code, 403)


class ConfiguracoesEmpresasTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.administrador = User.objects.create_user(
            username="admin_oeste",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = User.objects.create_user(
            username="cliente_comum",
            password="senha-segura",
        )

    def test_administrador_visualiza_configuracoes_na_tela_inicial(self):
        self.client.force_login(self.administrador)
        response = self.client.get(reverse("empresas:lista"))

        self.assertContains(response, "Configurações")
        self.assertContains(response, reverse("empresas:cadastrar"))
        self.assertNotContains(response, "Nova empresa")
        self.assertNotContains(response, "Parâmetros")

    def test_cliente_nao_acessa_configuracoes_diretamente(self):
        self.client.force_login(self.cliente)

        response = self.client.get(reverse("empresas:configuracoes"))

        self.assertEqual(response.status_code, 403)

    def test_administrador_cadastra_empresa(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse("empresas:cadastrar"),
            {
                "nome_fantasia": "Cliente Oeste",
                "nome": "Cliente Oeste Comércio Ltda",
                "cnpj": "12.345.678/0001-90",
                "grupo": "Grupo Oeste",
                "ativa": "on",
            },
        )

        self.assertRedirects(response, reverse("empresas:configuracoes"))
        empresa = Empresa.objects.get(cnpj="12.345.678/0001-90")
        self.assertEqual(empresa.nome_fantasia, "Cliente Oeste")
        self.assertEqual(empresa.grupo, "Grupo Oeste")
        self.assertTrue(empresa.ativa)

    def test_cadastro_rejeita_cnpj_incompleto(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse("empresas:cadastrar"),
            {
                "nome_fantasia": "Empresa Inválida",
                "nome": "Empresa Inválida Ltda",
                "cnpj": "123",
                "ativa": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Informe um CNPJ com 14 números.")
        self.assertFalse(Empresa.objects.filter(nome_fantasia="Empresa Inválida").exists())


class ParametrosOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_parametros",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_parametros",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa OMIE Ltda",
            nome_fantasia="Empresa OMIE",
            cnpj="00.000.000/0001-03",
        )

    def test_cliente_nao_acessa_parametros(self):
        self.client.force_login(self.cliente)
        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )
        self.assertEqual(response.status_code, 403)

    def test_administrador_seleciona_empresa(self):
        self.client.force_login(self.administrador)
        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )
        self.assertContains(response, "Credenciais exclusivas de")
        self.assertContains(response, self.empresa.nome_fantasia)

    def test_salva_credenciais_isoladas_e_criptografadas(self):
        self.client.force_login(self.administrador)
        response = self.client.post(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "app_key": "minha-app-key",
                "app_secret": "segredo-super-secreto",
                "ativa": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        integracao = IntegracaoOmie.objects.get(empresa=self.empresa)
        self.assertEqual(integracao.app_key, "minha-app-key")
        self.assertNotEqual(
            integracao.app_secret_criptografado,
            "segredo-super-secreto",
        )
        self.assertEqual(integracao.obter_app_secret(), "segredo-super-secreto")

    def test_edicao_sem_novo_secret_preserva_o_existente(self):
        integracao = IntegracaoOmie(empresa=self.empresa, app_key="key-antiga")
        integracao.definir_app_secret("secret-original")
        integracao.save()
        self.client.force_login(self.administrador)

        self.client.post(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "app_key": "key-nova",
                "app_secret": "",
                "ativa": "on",
            },
        )

        integracao.refresh_from_db()
        self.assertEqual(integracao.app_key, "key-nova")
        self.assertEqual(integracao.obter_app_secret(), "secret-original")


class EstruturaDRETests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_dre",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_dre",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa DRE Ltda",
            nome_fantasia="Empresa DRE",
            cnpj="00.000.000/0001-05",
        )
        self.url = reverse(
            "dashboards:dre_categorias",
            kwargs={"empresa_slug": self.empresa.slug},
        )

    def test_cards_dre_e_categorias_aparecem_separados_em_parametros(self):
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertContains(response, "<h2>DRE</h2>", html=True)
        self.assertContains(response, "<h2>Categorias</h2>", html=True)
        self.assertNotContains(response, "DRE e Categorias")
        self.assertContains(response, self.url)
        self.assertContains(
            response,
            reverse(
                "dashboards:categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )
        self.assertContains(
            response,
            reverse(
                "dashboards:metas",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )

    def test_metas_lista_vendedores_e_salva_valores(self):
        vendedor = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=101,
            nome="Ana Comercial",
        )
        VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=102,
            nome="Vendedor inativo",
            inativo=True,
        )
        url = reverse("dashboards:metas", kwargs={"empresa_slug": self.empresa.slug})
        self.client.force_login(self.administrador)

        response = self.client.get(url)

        self.assertContains(response, "Ana Comercial")
        self.assertNotContains(response, "Vendedor inativo")

        response = self.client.post(
            url,
            {
                "ano": "2026",
                "mes": "4",
                f"meta_{vendedor.pk}": "12.500,50",
                "acao": "salvar",
            },
        )

        self.assertRedirects(response, f"{url}?ano=2026&mes=4")
        meta = MetaVendedorComercial.objects.get(vendedor=vendedor)
        self.assertEqual(meta.valor_mensal, Decimal("12500.50"))
        self.assertEqual(meta.ano, 2026)
        self.assertEqual(meta.mes, 4)

    def test_metas_replica_valores_do_mes_ate_dezembro(self):
        vendedor = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=101,
            nome="Ana Comercial",
        )
        MetaVendedorComercial.objects.create(
            empresa=self.empresa,
            vendedor=vendedor,
            ano=2026,
            mes=3,
            valor_mensal=Decimal("3000"),
        )
        url = reverse("dashboards:metas", kwargs={"empresa_slug": self.empresa.slug})
        self.client.force_login(self.administrador)

        response = self.client.post(
            url,
            {
                "ano": "2026",
                "mes": "4",
                f"meta_{vendedor.pk}": "12.500,50",
                "acao": "replicar",
            },
        )

        self.assertRedirects(response, f"{url}?ano=2026&mes=4")
        self.assertEqual(
            MetaVendedorComercial.objects.get(
                vendedor=vendedor,
                ano=2026,
                mes=3,
            ).valor_mensal,
            Decimal("3000"),
        )
        for mes in range(4, 13):
            self.assertEqual(
                MetaVendedorComercial.objects.get(
                    vendedor=vendedor,
                    ano=2026,
                    mes=mes,
                ).valor_mensal,
                Decimal("12500.50"),
            )

    def test_usuario_comum_nao_acessa_metas(self):
        self.client.force_login(self.cliente)

        response = self.client.get(
            reverse("dashboards:metas", kwargs={"empresa_slug": self.empresa.slug})
        )

        self.assertEqual(response.status_code, 403)

    def test_usuario_comum_nao_acessa_estrutura_dre(self):
        self.client.force_login(self.cliente)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_cria_conta_pai_e_conta_filha(self):
        self.client.force_login(self.administrador)
        self.client.post(
            self.url,
            {
                "nome": "Receitas operacionais",
                "tipo": "pai",
                "conta_pai": "",
                "sinal": "+",
            },
        )
        pai = ContaDRE.objects.get(nome="Receitas operacionais")

        response = self.client.post(
            self.url,
            {
                "nome": "Receita de vendas",
                "tipo": "filho",
                "conta_pai": pai.pk,
                "sinal": "+",
            },
        )

        self.assertRedirects(response, self.url)
        filha = ContaDRE.objects.get(nome="Receita de vendas")
        self.assertEqual(filha.conta_pai, pai)
        self.assertEqual(filha.nivel, 2)

    def test_conta_filha_exige_conta_pai(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                "nome": "Conta sem grupo",
                "tipo": "filho",
                "conta_pai": "",
                "sinal": "-",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selecione a conta pai.")
        self.assertFalse(ContaDRE.objects.filter(nome="Conta sem grupo").exists())

    def test_cria_conta_de_resultado_sem_filhos(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                "nome": "Receita liquida",
                "tipo": "resultado",
                "conta_pai": "",
                "sinal": "+",
            },
        )

        self.assertRedirects(response, self.url)
        resultado = ContaDRE.objects.get(nome="Receita liquida")
        self.assertIsNone(resultado.conta_pai)
        self.assertEqual(resultado.sinal, ContaDRE.Sinal.RESULTADO)
        self.assertTrue(resultado.eh_resultado)

    def test_conta_de_resultado_nao_pode_receber_filha(self):
        resultado = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita liquida",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                "nome": "Conta indevida",
                "tipo": "filho",
                "conta_pai": resultado.pk,
                "sinal": "+",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ContaDRE.objects.filter(nome="Conta indevida").exists())

    def test_reordena_grupos_da_arvore(self):
        primeira = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Primeira",
            sinal="+",
            ordem=1,
        )
        segunda = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Segunda",
            sinal="=",
            ordem=2,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:reordenar_contas_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            data=json.dumps(
                {
                    "parent_id": None,
                    "ids": [segunda.pk, primeira.pk],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        primeira.refresh_from_db()
        segunda.refresh_from_db()
        self.assertEqual(segunda.ordem, 1)
        self.assertEqual(primeira.ordem, 2)

    def test_nao_exclui_grupo_com_conta_filha(self):
        pai = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Grupo",
            ordem=1,
        )
        ContaDRE.objects.create(
            empresa=self.empresa,
            conta_pai=pai,
            nome="Filha",
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:excluir_conta_dre",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "conta_id": pai.pk,
                },
            )
        )

        self.assertRedirects(response, self.url)
        self.assertTrue(ContaDRE.objects.filter(pk=pai.pk).exists())

    def test_exporta_planilha_dre_com_arvore_e_listas_de_selecao(self):
        pai = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receitas",
            sinal="+",
            ordem=1,
        )
        ContaDRE.objects.create(
            empresa=self.empresa,
            conta_pai=pai,
            nome="Vendas",
            sinal="-",
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:exportar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["DRE"]
        self.assertEqual(
            [ws.cell(1, coluna).value for coluna in range(1, 4)],
            ["Nome da conta DRE", "Tipo", "Operação"],
        )
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value], ["Receitas", "Pai", "+"])
        self.assertEqual([ws["A3"].value, ws["B3"].value, ws["C3"].value], ["Vendas", "Filho", "-"])
        self.assertTrue(ws.data_validations.dataValidation)

    def test_exporta_conta_de_resultado_com_tipo_proprio(self):
        ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita liquida",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:exportar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["DRE"]
        self.assertEqual(
            [ws["A2"].value, ws["B2"].value, ws["C2"].value],
            ["Receita liquida", "Resultado", "="],
        )

    def test_importa_planilha_dre_e_monta_relacao_pai_filho(self):
        self.client.force_login(self.administrador)
        planilha = arquivo_xlsx(
            "DRE",
            ("Nome da conta DRE", "Tipo", "Operação"),
            (
                ("Receitas", "Pai", "+"),
                ("Venda de serviços", "Filho", "+"),
                ("Resultado", "Pai", "="),
            ),
        )

        response = self.client.post(
            reverse(
                "dashboards:importar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {"planilha": planilha},
        )

        self.assertRedirects(response, self.url)
        receitas = ContaDRE.objects.get(empresa=self.empresa, nome="Receitas")
        venda = ContaDRE.objects.get(empresa=self.empresa, nome="Venda de serviços")
        self.assertEqual(venda.conta_pai, receitas)
        self.assertEqual(venda.ordem, 1)
        self.assertEqual(
            ContaDRE.objects.get(empresa=self.empresa, nome="Resultado").sinal,
            "=",
        )
        self.assertTrue(
            ContaDRE.objects.get(empresa=self.empresa, nome="Resultado").eh_resultado
        )

    def test_importacao_dre_rejeita_resultado_como_filho(self):
        self.client.force_login(self.administrador)
        planilha = arquivo_xlsx(
            "DRE",
            ("Nome da conta DRE", "Tipo", "OperaÃ§Ã£o"),
            (
                ("Receitas", "Pai", "+"),
                ("Receita liquida", "Filho", "="),
            ),
        )

        response = self.client.post(
            reverse(
                "dashboards:importar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {"planilha": planilha},
        )

        self.assertRedirects(response, self.url)
        self.assertFalse(ContaDRE.objects.filter(nome="Receita liquida").exists())

    def test_importacao_dre_exige_confirmacao_para_sobrepor(self):
        atual = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Estrutura atual",
            ordem=1,
        )
        self.client.force_login(self.administrador)
        url = reverse(
            "dashboards:importar_planilha_dre",
            kwargs={"empresa_slug": self.empresa.slug},
        )

        response = self.client.post(
            url,
            {
                "planilha": arquivo_xlsx(
                    "DRE",
                    ("Nome da conta DRE", "Tipo", "Operação"),
                    (("Nova estrutura", "Pai", "+"),),
                )
            },
        )

        self.assertRedirects(response, self.url)
        self.assertTrue(ContaDRE.objects.filter(pk=atual.pk).exists())
        self.assertFalse(ContaDRE.objects.filter(nome="Nova estrutura").exists())

        response = self.client.post(
            url,
            {
                "sobrepor": "sim",
                "planilha": arquivo_xlsx(
                    "DRE",
                    ("Nome da conta DRE", "Tipo", "Operação"),
                    (("Nova estrutura", "Pai", "+"),),
                ),
            },
        )

        self.assertRedirects(response, self.url)
        self.assertFalse(ContaDRE.objects.filter(pk=atual.pk).exists())
        self.assertTrue(ContaDRE.objects.filter(nome="Nova estrutura").exists())


class CategoriasOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_categorias",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_categorias",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa Categorias Ltda",
            nome_fantasia="Empresa Categorias",
            cnpj="00.000.000/0001-06",
        )
        self.url = reverse(
            "dashboards:categorias",
            kwargs={"empresa_slug": self.empresa.slug},
        )
        self.conta_dre = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita operacional",
            ordem=1,
        )
        self.categoria_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01",
            descricao="Receitas",
            conta_receita=True,
        )
        self.categoria_filha = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.01",
            categoria_superior="1.01",
            descricao="Venda de mercadorias",
            conta_receita=False,
        )
        self.categoria_inativa = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.02",
            categoria_superior="1.01",
            descricao="Categoria inativa",
            conta_inativa=True,
        )

    def test_exibe_apenas_ativas_com_regras_visuais_e_seletor_na_filha(self):
        self.client.force_login(self.administrador)

        response = self.client.get(self.url)

        self.assertContains(response, self.categoria_pai.codigo)
        self.assertContains(response, self.categoria_filha.codigo)
        self.assertNotContains(response, self.categoria_inativa.descricao)
        self.assertContains(response, "is-revenue")
        self.assertContains(response, "is-expense")
        self.assertContains(response, "is-parent")
        self.assertContains(response, "Salvar associações", count=2)
        self.assertNotContains(
            response,
            f'name="conta_dre_{self.categoria_pai.pk}"',
        )
        self.assertContains(
            response,
            f'name="conta_dre_{self.categoria_filha.pk}"',
        )

    def test_salva_vinculo_da_categoria_filha_com_conta_dre(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                f"conta_dre_{self.categoria_filha.pk}": self.conta_dre.pk,
            },
        )

        self.assertRedirects(response, self.url)
        self.categoria_filha.refresh_from_db()
        self.assertEqual(self.categoria_filha.conta_dre, self.conta_dre)

    def test_nao_oferece_conta_de_resultado_para_categoria(self):
        resultado = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita liquida",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=2,
        )
        self.client.force_login(self.administrador)

        response = self.client.get(self.url)

        self.assertContains(response, self.conta_dre.nome)
        self.assertNotContains(response, f'value="{resultado.pk}"')

    def test_rejeita_vinculo_de_categoria_com_conta_de_resultado(self):
        resultado = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita liquida",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=2,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                f"conta_dre_{self.categoria_filha.pk}": resultado.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.categoria_filha.refresh_from_db()
        self.assertIsNone(self.categoria_filha.conta_dre)

    def test_rejeita_conta_dre_de_outra_empresa(self):
        outra_empresa = Empresa.objects.create(
            nome="Outra Empresa Categorias Ltda",
            nome_fantasia="Outra Empresa Categorias",
            cnpj="00.000.000/0001-07",
        )
        conta_externa = ContaDRE.objects.create(
            empresa=outra_empresa,
            nome="Conta externa",
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                f"conta_dre_{self.categoria_filha.pk}": conta_externa.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uma das contas DRE selecionadas é inválida.")
        self.categoria_filha.refresh_from_db()
        self.assertIsNone(self.categoria_filha.conta_dre)

    def test_usuario_comum_nao_acessa_categorias(self):
        self.client.force_login(self.cliente)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_exporta_categorias_com_contas_dre_para_selecao(self):
        self.categoria_filha.conta_dre = self.conta_dre
        self.categoria_filha.save(update_fields=["conta_dre"])
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:exportar_planilha_categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["Categorias"]
        self.assertEqual([ws["A1"].value, ws["B1"].value], ["Categoria", "Conta DRE"])
        self.assertEqual(
            ws["A2"].value,
            f"{self.categoria_pai.codigo} - {self.categoria_pai.descricao}",
        )
        self.assertEqual(ws["B3"].value, self.conta_dre.nome)
        self.assertTrue(ws.data_validations.dataValidation)

    def test_importa_associacoes_de_categorias_pela_planilha(self):
        self.client.force_login(self.administrador)
        planilha = arquivo_xlsx(
            "Categorias",
            ("Categoria", "Conta DRE"),
            (
                (
                    f"{self.categoria_pai.codigo} - {self.categoria_pai.descricao}",
                    "",
                ),
                (
                    f"{self.categoria_filha.codigo} - {self.categoria_filha.descricao}",
                    self.conta_dre.nome,
                ),
            ),
        )

        response = self.client.post(
            reverse(
                "dashboards:importar_planilha_categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {"planilha": planilha},
        )

        self.assertRedirects(response, self.url)
        self.categoria_filha.refresh_from_db()
        self.assertEqual(self.categoria_filha.conta_dre, self.conta_dre)


class SincronizacaoClientesOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_sync_omie",
            password="senha-segura",
            is_staff=True,
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa Sync OMIE Ltda",
            nome_fantasia="Empresa Sync OMIE",
            cnpj="00.000.000/0001-04",
        )
        self.integracao = IntegracaoOmie(
            empresa=self.empresa,
            app_key="app-key",
        )
        self.integracao.definir_app_secret("app-secret")
        self.integracao.save()

    @patch("apps.empresas.views.iniciar_sincronizacao_omie")
    def test_endpoint_inicia_sincronizacao(self, iniciar_mock):
        self.client.force_login(self.administrador)
        response = self.client.post(
            reverse(
                "dashboards:sincronizar_clientes_omie",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 202)
        sincronizacao = SincronizacaoOmie.objects.get(empresa=self.empresa)
        iniciar_mock.assert_called_once_with(sincronizacao.pk)
        self.assertEqual(response.json()["status"], SincronizacaoOmie.Status.PENDENTE)

    @patch("apps.empresas.views.iniciar_sincronizacao_omie")
    def test_nova_sincronizacao_substitui_pendente_obsoleta(self, iniciar_mock):
        obsoleta = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            mensagem="Sincronização abandonada",
        )
        SincronizacaoOmie.objects.filter(pk=obsoleta.pk).update(
            atualizada_em=timezone.now() - timedelta(hours=1)
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:sincronizar_clientes_omie",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 202)
        obsoleta.refresh_from_db()
        self.assertEqual(obsoleta.status, SincronizacaoOmie.Status.ERRO)
        nova = SincronizacaoOmie.objects.exclude(pk=obsoleta.pk).get()
        iniciar_mock.assert_called_once_with(nova.pk)
        self.assertEqual(response.json()["id"], nova.pk)

    def test_polling_encerra_sincronizacao_obsoleta(self):
        obsoleta = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            status=SincronizacaoOmie.Status.EM_ANDAMENTO,
            mensagem="Processando",
        )
        SincronizacaoOmie.objects.filter(pk=obsoleta.pk).update(
            atualizada_em=timezone.now() - timedelta(hours=1)
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:status_sincronizacao_omie",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "sincronizacao_id": obsoleta.pk,
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], SincronizacaoOmie.Status.ERRO)
        self.assertIn("mais de 30 minutos", response.json()["erro"])

    @patch("apps.empresas.views._encerrar_sincronizacoes_omie_obsoletas")
    def test_polling_de_sincronizacao_ativa_nao_escreve_no_banco(
        self,
        encerrar_mock,
    ):
        ativa = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            status=SincronizacaoOmie.Status.EM_ANDAMENTO,
            mensagem="Processando",
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:status_sincronizacao_omie",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "sincronizacao_id": ativa.pk,
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()["status"],
            SincronizacaoOmie.Status.EM_ANDAMENTO,
        )
        encerrar_mock.assert_not_called()

    @patch("apps.empresas.omie.urlopen")
    def test_consultas_novas_usam_os_contratos_da_omie(self, urlopen_mock):
        urlopen_mock.return_value.__enter__.return_value.read.return_value = (
            b'{"pagina": 1, "total_de_paginas": 1, "total_de_registros": 0}'
        )

        consultar_tipos_conta_corrente(self.integracao, 2)
        consultar_contas_correntes(self.integracao, 3)
        consultar_contas_pagar(self.integracao, 4)
        consultar_contas_receber(self.integracao, 5)
        consultar_lancamentos_conta_corrente(self.integracao, 6)
        consultar_movimentos_financeiros(self.integracao, 7)
        consultar_produtos(self.integracao, 8)
        consultar_posicoes_estoque(self.integracao, 9)
        consultar_pedidos(self.integracao, 10)
        consultar_servicos(self.integracao, 11)
        consultar_ordens_servico(self.integracao, 12)
        consultar_contratos(self.integracao, 13)
        consultar_vendedores(self.integracao, 14)
        consultar_resumo_financas(self.integracao)
        conta_extrato = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=7090178721,
            codigo_integracao="",
            descricao="Conta extrato",
        )
        consultar_extrato_conta_corrente(
            self.integracao,
            conta_extrato,
            "01/01/2026",
            "23/07/2026",
        )

        requisicao_tipos = urlopen_mock.call_args_list[0].args[0]
        payload_tipos = json.loads(requisicao_tipos.data)
        self.assertTrue(requisicao_tipos.full_url.endswith("/geral/tipocc/"))
        self.assertEqual(payload_tipos["call"], "ListarTiposCC")
        self.assertEqual(
            payload_tipos["param"][0],
            {"pagina": 2, "registros_por_pagina": 50},
        )

        requisicao_contas = urlopen_mock.call_args_list[1].args[0]
        payload_contas = json.loads(requisicao_contas.data)
        self.assertEqual(
            requisicao_contas.full_url.split("?")[0].rstrip("/").split("/")[-1],
            "contacorrente",
        )
        self.assertEqual(payload_contas["call"], "ListarContasCorrentes")
        self.assertEqual(
            payload_contas["param"][0],
            {
                "pagina": 3,
                "registros_por_pagina": 100,
                "apenas_importado_api": "N",
            },
        )

        requisicao_contas_pagar = urlopen_mock.call_args_list[2].args[0]
        payload_contas_pagar = json.loads(requisicao_contas_pagar.data)
        self.assertTrue(
            requisicao_contas_pagar.full_url.endswith("/financas/contapagar/")
        )
        self.assertEqual(payload_contas_pagar["call"], "ListarContasPagar")
        self.assertEqual(
            payload_contas_pagar["param"][0],
            {
                "pagina": 4,
                "registros_por_pagina": 20,
                "apenas_importado_api": "N",
            },
        )

        requisicao_contas_receber = urlopen_mock.call_args_list[3].args[0]
        payload_contas_receber = json.loads(requisicao_contas_receber.data)
        self.assertTrue(
            requisicao_contas_receber.full_url.endswith(
                "/financas/contareceber/"
            )
        )
        self.assertEqual(payload_contas_receber["call"], "ListarContasReceber")
        self.assertEqual(
            payload_contas_receber["param"][0],
            {
                "pagina": 5,
                "registros_por_pagina": 20,
                "apenas_importado_api": "N",
            },
        )

        requisicao_lancamentos = urlopen_mock.call_args_list[4].args[0]
        payload_lancamentos = json.loads(requisicao_lancamentos.data)
        self.assertTrue(
            requisicao_lancamentos.full_url.endswith(
                "/financas/contacorrentelancamentos/"
            )
        )
        self.assertEqual(payload_lancamentos["call"], "ListarLancCC")
        self.assertEqual(
            payload_lancamentos["param"][0],
            {
                "nPagina": 6,
                "nRegPorPagina": 20,
            },
        )

        requisicao_movimentos = urlopen_mock.call_args_list[5].args[0]
        payload_movimentos = json.loads(requisicao_movimentos.data)
        self.assertTrue(requisicao_movimentos.full_url.endswith("/financas/mf/"))
        self.assertEqual(payload_movimentos["call"], "ListarMovimentos")
        self.assertEqual(
            payload_movimentos["param"][0],
            {
                "nPagina": 7,
                "nRegPorPagina": 500,
            },
        )

        requisicao_produtos = urlopen_mock.call_args_list[6].args[0]
        payload_produtos = json.loads(requisicao_produtos.data)
        self.assertTrue(requisicao_produtos.full_url.endswith("/geral/produtos/"))
        self.assertEqual(payload_produtos["call"], "ListarProdutos")
        self.assertEqual(
            payload_produtos["param"][0],
            {
                "pagina": 8,
                "registros_por_pagina": 50,
                "apenas_importado_api": "N",
                "filtrar_apenas_omiepdv": "N",
            },
        )

        requisicao_posicoes = urlopen_mock.call_args_list[7].args[0]
        payload_posicoes = json.loads(requisicao_posicoes.data)
        self.assertTrue(requisicao_posicoes.full_url.endswith("/estoque/consulta/"))
        self.assertEqual(payload_posicoes["call"], "ListarPosEstoque")
        self.assertTrue(payload_posicoes["param"][0].pop("dDataPosicao"))
        self.assertEqual(
            payload_posicoes["param"][0],
            {
                "nPagina": 9,
                "nRegPorPagina": 50,
                "cExibeTodos": "N",
                "codigo_local_estoque": 0,
            },
        )

        requisicao_pedidos = urlopen_mock.call_args_list[8].args[0]
        payload_pedidos = json.loads(requisicao_pedidos.data)
        self.assertTrue(requisicao_pedidos.full_url.endswith("/produtos/pedido/"))
        self.assertEqual(payload_pedidos["call"], "ListarPedidos")
        self.assertEqual(
            payload_pedidos["param"][0],
            {
                "pagina": 10,
                "registros_por_pagina": 100,
                "apenas_importado_api": "N",
            },
        )

        requisicao_servicos = urlopen_mock.call_args_list[9].args[0]
        payload_servicos = json.loads(requisicao_servicos.data)
        self.assertTrue(requisicao_servicos.full_url.endswith("/servicos/servico/"))
        self.assertEqual(payload_servicos["call"], "ListarCadastroServico")
        self.assertEqual(
            payload_servicos["param"][0],
            {
                "nPagina": 11,
                "nRegPorPagina": 20,
            },
        )

        requisicao_os = urlopen_mock.call_args_list[10].args[0]
        payload_os = json.loads(requisicao_os.data)
        self.assertTrue(requisicao_os.full_url.endswith("/servicos/os/"))
        self.assertEqual(payload_os["call"], "ListarOS")
        self.assertEqual(
            payload_os["param"][0],
            {
                "pagina": 12,
                "registros_por_pagina": 50,
                "apenas_importado_api": "N",
            },
        )

        requisicao_contratos = urlopen_mock.call_args_list[11].args[0]
        payload_contratos = json.loads(requisicao_contratos.data)
        self.assertTrue(
            requisicao_contratos.full_url.endswith("/servicos/contrato/")
        )
        self.assertEqual(payload_contratos["call"], "ListarContratos")
        self.assertEqual(
            payload_contratos["param"][0],
            {
                "pagina": 13,
                "registros_por_pagina": 50,
                "apenas_importado_api": "N",
            },
        )

        requisicao_vendedores = urlopen_mock.call_args_list[12].args[0]
        payload_vendedores = json.loads(requisicao_vendedores.data)
        self.assertTrue(requisicao_vendedores.full_url.endswith("/geral/vendedores/"))
        self.assertEqual(payload_vendedores["call"], "ListarVendedores")
        self.assertEqual(
            payload_vendedores["param"][0],
            {
                "pagina": 14,
                "registros_por_pagina": 100,
                "apenas_importado_api": "N",
            },
        )
        requisicao_resumo = urlopen_mock.call_args_list[13].args[0]
        payload_resumo = json.loads(requisicao_resumo.data)
        self.assertTrue(requisicao_resumo.full_url.endswith("/financas/resumo/"))
        self.assertEqual(payload_resumo["call"], "ObterResumoFinancas")
        self.assertTrue(payload_resumo["param"][0].pop("dDia"))
        self.assertEqual(
            payload_resumo["param"][0],
            {
                "lApenasResumo": True,
            },
        )
        requisicao_extrato = urlopen_mock.call_args_list[14].args[0]
        payload_extrato = json.loads(requisicao_extrato.data)
        self.assertTrue(requisicao_extrato.full_url.endswith("/financas/extrato/"))
        self.assertEqual(payload_extrato["call"], "ListarExtrato")
        self.assertEqual(
            payload_extrato["param"][0],
            {
                "nCodCC": 7090178721,
                "cCodIntCC": "",
                "dPeriodoInicial": "01/01/2026",
                "dPeriodoFinal": "23/07/2026",
                "cExibirApenasSaldo": "S",
            },
        )

    @override_settings(OMIE_API_RETRIES=2, OMIE_API_RETRY_DELAY=0)
    @patch("apps.empresas.omie.urlopen")
    def test_consulta_omie_tenta_novamente_quando_a_leitura_expira(self, urlopen_mock):
        resposta = urlopen_mock.return_value
        resposta.__enter__.return_value.read.return_value = (
            b'{"pagina": 1, "total_de_paginas": 1, "total_de_registros": 0}'
        )
        urlopen_mock.side_effect = [URLError("The read operation timed out"), resposta]

        dados = consultar_clientes(self.integracao, 1)

        self.assertEqual(dados["pagina"], 1)
        self.assertEqual(urlopen_mock.call_count, 2)

    @patch("apps.empresas.omie.consultar_extrato_conta_corrente")
    @patch("apps.empresas.omie.consultar_resumo_financas")
    @patch("apps.empresas.omie.consultar_contratos")
    @patch("apps.empresas.omie.consultar_ordens_servico")
    @patch("apps.empresas.omie.consultar_servicos")
    @patch("apps.empresas.omie.consultar_pedidos")
    @patch("apps.empresas.omie.consultar_posicoes_estoque")
    @patch("apps.empresas.omie.consultar_produtos")
    @patch("apps.empresas.omie.consultar_lancamentos_conta_corrente")
    @patch("apps.empresas.omie.consultar_movimentos_financeiros")
    @patch("apps.empresas.omie.consultar_contas_receber")
    @patch("apps.empresas.omie.consultar_contas_pagar")
    @patch("apps.empresas.omie.consultar_contas_correntes")
    @patch("apps.empresas.omie.consultar_tipos_conta_corrente")
    @patch("apps.empresas.omie.consultar_categorias")
    @patch("apps.empresas.omie.consultar_vendedores")
    @patch("apps.empresas.omie.consultar_departamentos")
    @patch("apps.empresas.omie.consultar_projetos")
    @patch("apps.empresas.omie.consultar_clientes")
    def test_importa_dados_omie_incluindo_contas_correntes(
        self,
        consultar_clientes_mock,
        consultar_projetos_mock,
        consultar_departamentos_mock,
        consultar_vendedores_mock,
        consultar_categorias_mock,
        consultar_tipos_conta_corrente_mock,
        consultar_contas_correntes_mock,
        consultar_contas_pagar_mock,
        consultar_contas_receber_mock,
        consultar_movimentos_financeiros_mock,
        consultar_lancamentos_conta_corrente_mock,
        consultar_produtos_mock,
        consultar_posicoes_estoque_mock,
        consultar_pedidos_mock,
        consultar_servicos_mock,
        consultar_ordens_servico_mock,
        consultar_contratos_mock,
        consultar_resumo_financas_mock,
        consultar_extrato_conta_corrente_mock,
    ):
        consultar_clientes_mock.side_effect = [
            {
                "pagina": 1,
                "total_de_paginas": 2,
                "total_de_registros": 2,
                "clientes_cadastro": [
                    {
                        "codigo_cliente_omie": 101,
                        "razao_social": "Cliente Um Ltda",
                        "nome_fantasia": "Cliente Um",
                        "cnpj_cpf": "00.000.000/0001-11",
                        "pessoa_fisica": "N",
                        "inativo": "N",
                        "tags": [{"tag": "Cliente"}],
                        "dadosBancarios": {"agencia": "1234"},
                    }
                ],
            },
            {
                "pagina": 2,
                "total_de_paginas": 2,
                "total_de_registros": 2,
                "clientes_cadastro": [
                    {
                        "codigo_cliente_omie": 202,
                        "razao_social": "Fornecedor Dois Ltda",
                        "cnpj_cpf": "00.000.000/0001-22",
                        "tags": [{"tag": "Fornecedor"}],
                    }
                ],
            },
        ]
        consultar_projetos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "cadastro": [
                {
                    "codigo": 303,
                    "codInt": "PROJ-303",
                    "nome": "Projeto 303",
                    "inativo": "N",
                    "info": {"data_alt": "01/04/2025"},
                }
            ],
        }
        consultar_departamentos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "departamentos": [
                {
                    "codigo": "5476993662",
                    "descricao": "Hinfoluz",
                    "estrutura": "001.001.001",
                    "inativo": "N",
                    "nivel_totalizador": "N",
                }
            ],
        }
        consultar_vendedores_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "cadastro": [
                {
                    "codInt": "CRM Omie",
                    "codigo": 4290841436,
                    "comissao": 0,
                    "email": "",
                    "fatura_pedido": "N",
                    "inativo": "N",
                    "nome": "CRM Omie",
                    "visualiza_pedido": "N",
                }
            ],
        }
        consultar_categorias_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "categoria_cadastro": [
                {
                    "categoria_superior": "0",
                    "codigo": "0.01",
                    "codigo_dre": "",
                    "conta_despesa": "N",
                    "conta_inativa": "N",
                    "conta_receita": "N",
                    "dadosDRE": {},
                    "definida_pelo_usuario": "N",
                    "descricao": "Transferência",
                    "descricao_padrao": "Transferência",
                    "id_conta_contabil": "",
                    "nao_exibir": "S",
                    "natureza": "",
                    "tag_conta_contabil": "",
                    "tipo_categoria": "",
                    "totalizadora": "S",
                    "transferencia": "S",
                }
            ],
        }
        consultar_produtos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "produto_servico_cadastro": [
                {
                    "bloqueado": "N",
                    "bloquear_exclusao": "N",
                    "codInt_familia": "",
                    "codigo": "PRD00041",
                    "codigo_familia": 0,
                    "codigo_produto": 3293025013,
                    "codigo_produto_integracao": "",
                    "descr_detalhada": "",
                    "descricao": "FILME DE POLIETILENO",
                    "descricao_familia": "",
                    "ean": "",
                    "estoque_minimo": 0,
                    "importado_api": "N",
                    "inativo": "N",
                    "info": {
                        "dAlt": "02/07/2021",
                        "dInc": "02/07/2021",
                    },
                    "marca": "",
                    "modelo": "",
                    "ncm": "3915.90.00",
                    "peso_bruto": 0,
                    "peso_liq": 0,
                    "produto_lote": "N",
                    "produto_variacao": "N",
                    "quantidade_estoque": 0,
                    "recomendacoes_fiscais": {
                        "cupom_fiscal": "N",
                        "market_place": "N",
                    },
                    "tipoItem": "99",
                    "unidade": "KG",
                    "valor_unitario": 0,
                }
            ],
        }
        consultar_posicoes_estoque_mock.return_value = {
            "nPagina": 1,
            "nTotPaginas": 1,
            "nRegistros": 1,
            "nTotRegistros": 1,
            "produtos": [
                {
                    "cCodInt": "",
                    "cCodigo": "PRD00041",
                    "cDescricao": "FILME DE POLIETILENO",
                    "codigo_local_estoque": 3036783070,
                    "estoque_minimo": 0,
                    "fisico": 10,
                    "nCMC": 1.25,
                    "nCodProd": 3293025013,
                    "nPendente": 2,
                    "nPrecoUnitario": 2.1,
                    "nSaldo": 8,
                    "reservado": 1,
                }
            ],
        }
        consultar_servicos_mock.return_value = {
            "nPagina": 1,
            "nTotPaginas": 1,
            "nRegistros": 1,
            "nTotRegistros": 1,
            "cadastros": [
                {
                    "cabecalho": {
                        "cCodCateg": "0.01",
                        "cCodLC116": "14.06",
                        "cCodServMun": "1702",
                        "cCodigo": "SRV00001",
                        "cDescricao": "SERVICO DE CALIBRACAO",
                        "cIdTrib": "01",
                        "cTipoDesc": "",
                        "nAliqDesc": 0,
                        "nPrecoUnit": 150,
                        "nValorDesc": 0,
                    },
                    "descricao": {
                        "cDescrCompleta": "SERVICO DE CALIBRACAO",
                    },
                    "impostos": {
                        "cRetCOFINS": "N",
                        "cRetCSLL": "N",
                        "cRetINSS": "N",
                        "cRetIR": "N",
                        "cRetISS": "N",
                        "cRetPIS": "N",
                        "lDeduzISS": False,
                        "nAliqISS": 5,
                    },
                    "info": {
                        "cImpAPI": "N",
                        "dAlt": "30/09/2025",
                        "dInc": "30/09/2025",
                        "inativo": "N",
                    },
                    "intListar": {
                        "cCodIntServ": "",
                        "nCodServ": 4290828325,
                    },
                }
            ],
        }
        consultar_ordens_servico_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "osCadastro": [
                {
                    "Cabecalho": {
                        "cCodIntOS": "",
                        "cCodParc": "000",
                        "cEtapa": "60",
                        "cNumOS": "2",
                        "dDtPrevisao": "23/05/2026",
                        "nCodCli": 101,
                        "nCodOS": 4362685093,
                        "nCodVend": 4290841436,
                        "nQtdeParc": 1,
                        "nValorTotal": 200,
                        "nValorTotalImpRet": 0,
                    },
                    "Departamentos": [],
                    "Email": {
                        "cEnvBoleto": "N",
                        "cEnvLink": "N",
                        "cEnvPix": "N",
                        "cEnvRecibo": "S",
                        "cEnviarPara": "cliente@example.com",
                    },
                    "InfoCadastro": {
                        "cCancelada": "N",
                        "cFaturada": "S",
                        "cOrigem": "CTR",
                        "dDtAlt": "23/05/2026",
                        "dDtFat": "23/05/2026",
                        "dDtInc": "23/05/2026",
                    },
                    "InformacoesAdicionais": {
                        "cCidPrestServ": "ARACOIABA DA SERRA (SP)",
                        "cCodCateg": "0.01",
                        "cNumContrato": "2026/00001",
                        "cNumRecibo": "1",
                        "cUsoConsumo": "N",
                        "nCodCC": 3036783065,
                    },
                    "Observacoes": {},
                    "Parcelas": [
                        {
                            "dDtVenc": "10/06/2026",
                            "nDias": 0,
                            "nParcela": 1,
                            "nPercentual": 100,
                            "nValor": 200,
                        }
                    ],
                    "ServicosPrestados": [
                        {
                            "cCodCategItem": "0.01",
                            "cCodServLC116": "14.06",
                            "cCodServMun": "1702",
                            "cDescServ": "SERVICO DE CALIBRACAO",
                            "cNaoGerarFinanceiro": "N",
                            "cReembolso": "N",
                            "cRetemISS": "",
                            "cTribServ": "01",
                            "impostos": {
                                "cRetemCOFINS": "N",
                                "cRetemCSLL": "N",
                                "cRetemINSS": "N",
                                "cRetemIRRF": "N",
                                "cRetemPIS": "N",
                                "lDeduzISS": False,
                                "nAliqISS": 5,
                                "nBaseISS": 150,
                                "nValorISS": 7.5,
                            },
                            "nAliqDesconto": 0,
                            "nCodServico": 4290828325,
                            "nIdItem": 4362684824,
                            "nQtde": 1,
                            "nSeqItem": 1,
                            "nValUnit": 150,
                            "nValorAcrescimos": 0,
                            "nValorDesconto": 0,
                            "nValorOutrasRetencoes": 0,
                        },
                        {
                            "cCodCategItem": "0.01",
                            "cCodServLC116": "14.06",
                            "cCodServMun": "1702",
                            "cDescServ": "Despesas reembolsaveis",
                            "cNaoGerarFinanceiro": "N",
                            "cReembolso": "S",
                            "cRetemISS": "",
                            "cTribServ": "",
                            "impostos": {
                                "lDeduzISS": False,
                                "nAliqISS": 0,
                                "nBaseISS": 0,
                                "nValorISS": 0,
                            },
                            "nAliqDesconto": 0,
                            "nCodServico": 0,
                            "nIdItem": 4362685095,
                            "nQtde": 1,
                            "nSeqItem": 2,
                            "nValUnit": 50,
                            "nValorAcrescimos": 0,
                            "nValorDesconto": 0,
                            "nValorOutrasRetencoes": 0,
                        },
                    ],
                }
            ],
        }
        consultar_contratos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "contratoCadastro": [
                {
                    "cabecalho": {
                        "cCodIntCtr": "",
                        "cCodSit": "10",
                        "cNumCtr": "2026/00001",
                        "cTipoFat": "01",
                        "dVigFinal": "26/05/2029",
                        "dVigInicial": "05/03/2025",
                        "nCodCli": 101,
                        "nCodCtr": 4362684823,
                        "nDiaFat": 1,
                        "nValTotMes": 150,
                    },
                    "departamentos": [],
                    "despesasReembolsaveis": {
                        "cCodCategReemb": "0.01",
                        "despesaReembolsavel": [
                            {
                                "cDescReemb": "motop",
                                "cRecorrenteReemb": "N",
                                "dDataReemb": "23/05/2026",
                                "nCodReemb": 4362684835,
                                "nValorReemb": 50,
                            }
                        ],
                    },
                    "emailCliente": {
                        "cEnviarBoleto": "N",
                        "cEnviarLinkNfse": "N",
                        "cEnviarPix": "N",
                        "cEnviarRecibo": "S",
                    },
                    "infAdic": {
                        "cCidPrestServ": "ARACOIABA DA SERRA (SP)",
                        "cCodCateg": "0.01",
                        "cUsoConsumo": "N",
                        "nCodCC": 3036783065,
                        "nCodProj": 303,
                        "nCodVend": 0,
                    },
                    "itensContrato": [
                        {
                            "itemCabecalho": {
                                "aliqDesconto": 0,
                                "cCodCategItem": "0.01",
                                "cNaoGerarFinanceiro": "N",
                                "cTpDesconto": "",
                                "codItem": 4362684824,
                                "codLC116": "14.06",
                                "codNBS": "",
                                "codServMunic": "1702",
                                "codServico": 4290828325,
                                "natOperacao": "01",
                                "quant": 1,
                                "seq": 1,
                                "valorAcrescimo": 0,
                                "valorDed": 0,
                                "valorDesconto": 0,
                                "valorOutrasRetencoes": 0,
                                "valorTotal": 150,
                                "valorUnit": 150,
                            },
                            "itemDescrServ": {
                                "descrCompleta": "SERVICO DE CALIBRACAO",
                            },
                            "itemImpostos": {
                                "aliqISS": 5,
                                "lDeduzISS": False,
                                "retISS": "N",
                                "valorISS": 7.5,
                            },
                            "itemLeiTranspImp": {},
                        }
                    ],
                    "observacoes": {"cObsContrato": ""},
                    "vencTextos": {
                        "cCodPerRef": "001",
                        "cPostergar": "S",
                        "cProxMes": "S",
                        "cTpVenc": "002",
                        "nDiaFixo": 10,
                        "nDias": 5,
                    },
                }
            ],
        }
        consultar_tipos_conta_corrente_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "cadastros": [
                {
                    "cCodigo": "CX",
                    "cDescricao": "Caixinha",
                    "cGrupo": "CX",
                }
            ],
        }
        consultar_contas_correntes_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "ListarContasCorrentes": [
                {
                    "nCodCC": 3036783065,
                    "cCodCCInt": "CAIXA-01",
                    "tipo_conta_corrente": "CX",
                    "codigo_banco": "999",
                    "descricao": "Caixinha",
                    "codigo_agencia": "",
                    "numero_conta_corrente": "",
                    "saldo_inicial": 125.5,
                    "saldo_data": "08/10/2020",
                    "valor_limite": 500,
                    "nao_fluxo": "N",
                    "nao_resumo": "S",
                    "cobr_sn": "N",
                    "bol_sn": "N",
                    "pix_sn": "S",
                    "importado_api": "N",
                    "bloqueado": "N",
                    "inativo": "N",
                    "observacao": "Conta de caixa",
                }
            ],
        }
        consultar_resumo_financas_mock.return_value = {
            "dDia": "22/07/2026",
            "contaCorrente": {
                "cCor": "808080",
                "cIcone": "f472",
                "vLimiteCredito": 0,
                "vTotal": 219720.53,
            },
        }
        consultar_extrato_conta_corrente_mock.return_value = {
            "nCodCC": 3036783065,
            "nSaldoProvisorio": 9876.54,
        }
        consultar_contas_pagar_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "conta_pagar_cadastro": [
                {
                    "categorias": [
                        {
                            "codigo_categoria": "0.01",
                            "percentual": 100,
                            "valor": 74.89,
                        }
                    ],
                    "cnab_integracao_bancaria": {
                        "codigo_forma_pagamento": "PIX",
                        "pix_qrcode": "000201010212",
                    },
                    "codigo_categoria": "0.01",
                    "codigo_cliente_fornecedor": 202,
                    "codigo_lancamento_integracao": "",
                    "codigo_lancamento_omie": 1344090316,
                    "codigo_projeto": 303,
                    "codigo_tipo_documento": "BOL",
                    "data_emissao": "07/10/2025",
                    "data_entrada": "07/10/2025",
                    "data_previsao": "10/11/2025",
                    "data_vencimento": "04/11/2025",
                    "distribuicao": [],
                    "id_conta_corrente": 3036783065,
                    "id_origem": "APIP",
                    "info": {"cImpAPI": "N"},
                    "numero_documento": "RPS 3071844",
                    "numero_documento_fiscal": "3080816",
                    "numero_parcela": "001/001",
                    "retem_cofins": "N",
                    "retem_csll": "N",
                    "retem_inss": "N",
                    "retem_ir": "N",
                    "retem_iss": "N",
                    "retem_pis": "N",
                    "status_titulo": "ATRASADO",
                    "valor_documento": 74.89,
                }
            ],
        }
        consultar_contas_receber_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "conta_receber_cadastro": [
                {
                    "boleto": {
                        "cGerado": "N",
                        "cNumBancario": "",
                        "cNumBoleto": "",
                        "dDtEmBol": "",
                        "nPerJuros": 0,
                        "nPerMulta": 0,
                    },
                    "categorias": [
                        {
                            "codigo_categoria": "0.01",
                            "percentual": 100,
                            "valor": 500,
                        }
                    ],
                    "chave_nfe": "35201013835876000108550010000000841966893592",
                    "codigo_categoria": "0.01",
                    "codigo_cliente_fornecedor": 101,
                    "codigo_lancamento_integracao": "",
                    "codigo_lancamento_omie": 3037854759,
                    "codigo_projeto": 303,
                    "codigo_tipo_documento": "99999",
                    "data_emissao": "20/10/2020",
                    "data_previsao": "20/10/2020",
                    "data_registro": "20/10/2020",
                    "data_vencimento": "20/10/2020",
                    "distribuicao": [],
                    "id_conta_corrente": 3036783065,
                    "id_origem": "VENR",
                    "info": {"cImpAPI": "N"},
                    "nCodPedido": 3037850626,
                    "numero_documento_fiscal": "00000084",
                    "numero_parcela": "001/001",
                    "numero_pedido": "5",
                    "operacao": "11",
                    "retem_cofins": "N",
                    "retem_csll": "N",
                    "retem_inss": "N",
                    "retem_ir": "N",
                    "retem_iss": "N",
                    "retem_pis": "N",
                    "status_titulo": "CANCELADO",
                    "tipo_agrupamento": "I",
                    "valor_documento": 500,
                }
            ],
        }
        consultar_movimentos_financeiros_mock.return_value = {
            "nPagina": 1,
            "nTotPaginas": 1,
            "nRegistros": 1,
            "nTotRegistros": 1,
            "movimentos": [
                {
                    "detalhes": {
                        "cCPFCNPJCliente": "00.000.000/0002-22",
                        "cCodCateg": "0.01",
                        "cGrupo": "CONTA_A_PAGAR",
                        "cNatureza": "P",
                        "cNumBoleto": "479487196",
                        "cNumParcela": "001/001",
                        "cNumTitulo": "RPS 3071844",
                        "cOrigem": "APIP",
                        "cStatus": "PAGO",
                        "cTipo": "BOL",
                        "dDtEmissao": "07/10/2025",
                        "dDtPagamento": "15/12/2025",
                        "dDtPrevisao": "10/11/2025",
                        "dDtRegistro": "07/10/2025",
                        "dDtVenc": "04/11/2025",
                        "nCodCC": 3036783065,
                        "nCodCliente": 202,
                        "nCodProjeto": 303,
                        "nCodTitRepet": 1344090316,
                        "nCodTitulo": 1344090316,
                        "nValorTitulo": 10000,
                    },
                    "resumo": {
                        "cLiquidado": "N",
                        "nDesconto": 0,
                        "nJuros": 0,
                        "nMulta": 0,
                        "nValAberto": 5000,
                        "nValLiquido": 5000,
                        "nValPago": 5000,
                    },
                }
            ],
        }
        consultar_lancamentos_conta_corrente_mock.return_value = {
            "nPagina": 1,
            "nTotPaginas": 1,
            "nRegistros": 1,
            "nTotRegistros": 1,
            "listaLancamentos": [
                {
                    "cCodIntLanc": "",
                    "cabecalho": {
                        "dDtLanc": "08/04/2026",
                        "nCodCC": 3036783065,
                        "nValorLanc": 2289.5,
                    },
                    "departamentos": [
                        {
                            "cCodDep": "5476993662",
                            "nPerDep": 100,
                            "nValDep": 2289.5,
                        }
                    ],
                    "detalhes": {
                        "aCodCateg": [],
                        "cCodCateg": "0.01",
                        "cNumDoc": "",
                        "cObs": "",
                        "cTipo": "99999",
                        "nCodCliente": 101,
                        "nCodProjeto": 303,
                    },
                    "diversos": {
                        "cHrConc": "",
                        "cIdentLanc": "",
                        "cNatureza": "R",
                        "cOrigem": "BAXR",
                        "cUsConc": "",
                        "dDtConc": "",
                        "nCodComprador": 0,
                        "nCodLancCR": 3037854759,
                        "nCodVendedor": 0,
                    },
                    "info": {
                        "cImpAPI": "N",
                        "dAlt": "17/04/2026",
                        "dInc": "17/04/2026",
                    },
                    "nCodAgrup": 4353703652,
                    "nCodLanc": 4353703652,
                    "transferencia": {"nCodCCDestino": 0},
                }
            ],
        }
        consultar_pedidos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "pedido_venda_produto": [
                {
                    "cabecalho": {
                        "bloqueado": "N",
                        "codigo_cenario_impostos": "3037132865",
                        "codigo_cliente": 101,
                        "codigo_empresa": 3036783056,
                        "codigo_parcela": "000",
                        "codigo_pedido": 3037132866,
                        "codigo_pedido_integracao": "",
                        "data_previsao": "14/10/2020",
                        "encerrado": "",
                        "etapa": "60",
                        "importado_api": "",
                        "numero_pedido": "1",
                        "origem_pedido": "",
                        "qtde_parcelas": 0,
                        "quantidade_itens": 1,
                    },
                    "departamentos": [],
                    "det": [
                        {
                            "ide": {
                                "codigo_item": 3037133403,
                                "codigo_item_integracao": "",
                                "simples_nacional": "N",
                            },
                            "imposto": {"icms": {}},
                            "inf_adic": {
                                "codigo_categoria_item": "0.01",
                                "codigo_cenario_impostos_item": "3037132865",
                                "codigo_local_estoque": 3036783070,
                                "nao_gerar_financeiro": "S",
                                "nao_movimentar_estoque": "N",
                                "nao_somar_total": "N",
                            },
                            "produto": {
                                "cfop": "5.102",
                                "codigo": "PRD00041",
                                "codigo_produto": 3293025013,
                                "descricao": "FILME DE POLIETILENO",
                                "ncm": "3915.90.00",
                                "percentual_desconto": 0,
                                "quantidade": 3848,
                                "reservado": "N",
                                "unidade": "KG",
                                "valor_desconto": 0,
                                "valor_mercadoria": 8080.8,
                                "valor_total": 8080.8,
                                "valor_unitario": 2.1,
                            },
                        }
                    ],
                    "exportacao": {"nao_exportacao": "N"},
                    "frete": {
                        "modalidade": "9",
                        "valor_frete": 0,
                        "valor_seguro": 0,
                    },
                    "infoCadastro": {
                        "autorizado": "S",
                        "cImpAPI": "N",
                        "cancelado": "N",
                        "dAlt": "14/10/2020",
                        "dFat": "14/10/2020",
                        "dInc": "14/10/2020",
                        "denegado": "N",
                        "devolvido": "N",
                        "devolvido_parcial": "N",
                        "faturado": "S",
                    },
                    "informacoes_adicionais": {
                        "codProj": 303,
                        "codVend": 0,
                        "codigo_categoria": "0.01",
                        "codigo_conta_corrente": 3036783065,
                        "consumidor_final": "N",
                    },
                    "lista_parcelas": {"parcela": []},
                    "observacoes": {"obs_venda": ""},
                    "total_pedido": {
                        "valor_descontos": 0,
                        "valor_mercadorias": 8080.8,
                        "valor_total_pedido": 8080.8,
                    },
                }
            ],
        }
        sincronizacao = SincronizacaoOmie.objects.create(empresa=self.empresa)

        executar_sincronizacao_omie(sincronizacao.pk)

        sincronizacao.refresh_from_db()
        self.assertEqual(sincronizacao.status, SincronizacaoOmie.Status.CONCLUIDA)
        self.assertEqual(sincronizacao.pagina_atual, 18)
        self.assertEqual(sincronizacao.registros_processados, 20)
        self.assertEqual(CadastroOmie.objects.count(), 2)
        self.assertEqual(
            CadastroOmie.objects.get(codigo_cliente_omie=101).tipo,
            CadastroOmie.Tipo.CLIENTE,
        )
        self.assertEqual(
            CadastroOmie.objects.get(codigo_cliente_omie=202).tipo,
            CadastroOmie.Tipo.FORNECEDOR,
        )
        projeto = ProjetoOmie.objects.get(codigo=303)
        self.assertEqual(projeto.nome, "Projeto 303")
        self.assertEqual(projeto.codigo_integracao, "PROJ-303")
        departamento = DepartamentoOmie.objects.get(codigo="5476993662")
        self.assertEqual(departamento.descricao, "Hinfoluz")
        self.assertEqual(departamento.estrutura, "001.001.001")
        vendedor = VendedorOmie.objects.get(codigo=4290841436)
        self.assertEqual(vendedor.nome, "CRM Omie")
        self.assertEqual(vendedor.codigo_integracao, "CRM Omie")
        self.assertEqual(str(vendedor.comissao), "0.0000")
        self.assertFalse(vendedor.fatura_pedido)
        self.assertFalse(vendedor.visualiza_pedido)
        self.assertFalse(vendedor.inativo)
        self.assertEqual(vendedor.dados_originais["nome"], "CRM Omie")
        produto = ProdutoOmie.objects.get(codigo_produto=3293025013)
        self.assertEqual(produto.codigo, "PRD00041")
        self.assertEqual(produto.descricao, "FILME DE POLIETILENO")
        self.assertEqual(produto.unidade, "KG")
        self.assertEqual(produto.ncm, "3915.90.00")
        self.assertFalse(produto.inativo)
        self.assertEqual(produto.info["dAlt"], "02/07/2021")
        posicao = PosicaoEstoqueOmie.objects.get(codigo_produto=3293025013)
        self.assertEqual(posicao.produto, produto)
        self.assertEqual(posicao.codigo, "PRD00041")
        self.assertEqual(posicao.codigo_local_estoque, 3036783070)
        self.assertEqual(str(posicao.cmc), "1.2500")
        self.assertEqual(str(posicao.saldo), "8.0000")
        self.assertEqual(str(posicao.fisico), "10.0000")
        self.assertEqual(posicao.dados_originais["nCMC"], 1.25)
        categoria = CategoriaOmie.objects.get(codigo="0.01")
        self.assertEqual(categoria.descricao, "Transferência")
        self.assertEqual(categoria.categoria_superior, "0")
        self.assertTrue(categoria.totalizadora)
        self.assertTrue(categoria.transferencia)
        self.assertTrue(categoria.nao_exibir)
        self.assertFalse(categoria.conta_inativa)
        servico = ServicoOmie.objects.get(codigo_servico=4290828325)
        self.assertEqual(servico.codigo, "SRV00001")
        self.assertEqual(servico.descricao, "SERVICO DE CALIBRACAO")
        self.assertEqual(servico.descricao_completa, "SERVICO DE CALIBRACAO")
        self.assertEqual(servico.categoria_principal, categoria)
        self.assertEqual(servico.codigo_lc116, "14.06")
        self.assertEqual(str(servico.preco_unitario), "150.0000")
        self.assertEqual(str(servico.aliquota_iss), "5.0000")
        self.assertFalse(servico.inativo)
        self.assertEqual(servico.info["dAlt"], "30/09/2025")
        tipo_conta = TipoContaCorrenteOmie.objects.get(codigo="CX")
        self.assertEqual(tipo_conta.descricao, "Caixinha")
        self.assertEqual(tipo_conta.grupo, "CX")
        conta = ContaCorrenteOmie.objects.get(codigo_omie=3036783065)
        self.assertEqual(conta.tipo_conta, tipo_conta)
        self.assertEqual(conta.tipo_codigo, "CX")
        self.assertEqual(conta.descricao, "Caixinha")
        self.assertEqual(str(conta.saldo_inicial), "125.50")
        self.assertEqual(str(conta.valor_limite), "500.00")
        self.assertTrue(conta.nao_resumo)
        self.assertTrue(conta.emite_pix)
        self.assertFalse(conta.inativo)
        self.assertEqual(str(conta.saldo_atual), "9876.54")
        self.assertIsNotNone(conta.saldo_atualizado_em)
        self.assertEqual(conta.dados_originais["extrato"]["nSaldoProvisorio"], 9876.54)
        consultar_extrato_conta_corrente_mock.assert_called_once()
        self.assertEqual(consultar_extrato_conta_corrente_mock.call_args.args[1], conta)
        self.assertEqual(conta.dados_originais["codigo_banco"], "999")
        self.empresa.refresh_from_db()
        self.assertEqual(str(self.empresa.saldo_contas_omie), "219720.53")
        self.assertIsNotNone(self.empresa.saldo_contas_atualizado_em)
        self.assertEqual(
            self.empresa.resumo_financeiro_omie["contaCorrente"]["vTotal"],
            219720.53,
        )
        contrato = ContratoOmie.objects.get(codigo_contrato=4362684823)
        self.assertEqual(contrato.numero_contrato, "2026/00001")
        self.assertEqual(contrato.cliente.codigo_cliente_omie, 101)
        self.assertEqual(contrato.conta_corrente, conta)
        self.assertEqual(contrato.categoria_principal, categoria)
        self.assertEqual(contrato.categoria_reembolso, categoria)
        self.assertEqual(contrato.projeto, projeto)
        self.assertEqual(contrato.vigencia_inicial.isoformat(), "2025-03-05")
        self.assertEqual(contrato.vigencia_final.isoformat(), "2029-05-26")
        self.assertEqual(str(contrato.valor_total_mes), "150.0000")
        self.assertEqual(
            contrato.despesas_reembolsaveis["despesaReembolsavel"][0]["nValorReemb"],
            50,
        )
        item_contrato = ContratoItemOmie.objects.get(codigo_item=4362684824)
        self.assertEqual(item_contrato.contrato, contrato)
        self.assertEqual(item_contrato.servico, servico)
        self.assertEqual(item_contrato.categoria_principal, categoria)
        self.assertEqual(str(item_contrato.quantidade), "1.0000")
        self.assertEqual(str(item_contrato.valor_unitario), "150.0000")
        self.assertEqual(str(item_contrato.valor_total), "150.0000")
        self.assertEqual(str(item_contrato.valor_iss), "7.5000")
        ordem_servico = OrdemServicoOmie.objects.get(codigo_os=4362685093)
        self.assertEqual(ordem_servico.numero_os, "2")
        self.assertEqual(ordem_servico.cliente.codigo_cliente_omie, 101)
        self.assertEqual(ordem_servico.conta_corrente, conta)
        self.assertEqual(ordem_servico.categoria_principal, categoria)
        self.assertEqual(ordem_servico.codigo_vendedor, 4290841436)
        self.assertTrue(ordem_servico.faturada)
        self.assertFalse(ordem_servico.cancelada)
        self.assertEqual(ordem_servico.data_previsao.isoformat(), "2026-05-23")
        self.assertEqual(str(ordem_servico.valor_total), "200.0000")
        self.assertEqual(len(ordem_servico.parcelas), 1)
        item_os = OrdemServicoItemOmie.objects.get(codigo_item=4362684824)
        self.assertEqual(item_os.ordem_servico, ordem_servico)
        self.assertEqual(item_os.servico, servico)
        self.assertEqual(item_os.categoria_principal, categoria)
        self.assertEqual(str(item_os.quantidade), "1.0000")
        self.assertEqual(str(item_os.valor_unitario), "150.0000")
        self.assertEqual(str(item_os.valor_iss), "7.5000")
        self.assertFalse(item_os.reembolso)
        reembolso = OrdemServicoItemOmie.objects.get(codigo_item=4362685095)
        self.assertIsNone(reembolso.servico)
        self.assertTrue(reembolso.reembolso)
        self.assertEqual(str(reembolso.valor_unitario), "50.0000")
        pedido = PedidoOmie.objects.get(codigo_pedido=3037132866)
        self.assertEqual(pedido.numero_pedido, "1")
        self.assertEqual(pedido.cliente.codigo_cliente_omie, 101)
        self.assertEqual(pedido.conta_corrente, conta)
        self.assertEqual(pedido.categoria_principal, categoria)
        self.assertEqual(pedido.projeto, projeto)
        self.assertTrue(pedido.faturado)
        self.assertFalse(pedido.cancelado)
        self.assertEqual(pedido.data_previsao.isoformat(), "2020-10-14")
        self.assertEqual(str(pedido.valor_total_pedido), "8080.8000")
        item_pedido = PedidoItemOmie.objects.get(codigo_item=3037133403)
        self.assertEqual(item_pedido.pedido, pedido)
        self.assertEqual(item_pedido.produto, produto)
        self.assertEqual(item_pedido.categoria_principal, categoria)
        self.assertEqual(str(item_pedido.quantidade), "3848.0000")
        self.assertEqual(str(item_pedido.valor_unitario), "2.1000")
        self.assertTrue(item_pedido.nao_gerar_financeiro)
        self.assertFalse(item_pedido.nao_movimentar_estoque)
        conta_pagar = ContaPagarOmie.objects.get(
            codigo_lancamento_omie=1344090316
        )
        self.assertEqual(conta_pagar.fornecedor.codigo_cliente_omie, 202)
        self.assertEqual(conta_pagar.conta_corrente, conta)
        self.assertEqual(conta_pagar.categoria_principal, categoria)
        self.assertEqual(conta_pagar.projeto, projeto)
        self.assertEqual(conta_pagar.data_vencimento.isoformat(), "2025-11-04")
        self.assertEqual(str(conta_pagar.valor_documento), "74.89")
        self.assertEqual(str(conta_pagar.valor_a_pagar), "74.89")
        self.assertEqual(conta_pagar.status_titulo, "ATRASADO")
        self.assertEqual(
            conta_pagar.cnab_integracao_bancaria["codigo_forma_pagamento"],
            "PIX",
        )
        movimento = MovimentoFinanceiroOmie.objects.get(
            codigo_titulo=1344090316
        )
        self.assertEqual(movimento.conta_pagar, conta_pagar)
        self.assertEqual(movimento.cliente_fornecedor.codigo_cliente_omie, 202)
        self.assertEqual(movimento.conta_corrente, conta)
        self.assertEqual(movimento.categoria_principal, categoria)
        self.assertEqual(movimento.projeto, projeto)
        self.assertEqual(movimento.grupo, "CONTA_A_PAGAR")
        self.assertEqual(movimento.natureza, "P")
        self.assertEqual(movimento.status, "PAGO")
        self.assertFalse(movimento.liquidado)
        self.assertEqual(str(movimento.valor_titulo), "10000.00")
        self.assertEqual(str(movimento.valor_aberto), "5000.00")
        self.assertEqual(str(movimento.valor_pago), "5000.00")
        conta_receber = ContaReceberOmie.objects.get(
            codigo_lancamento_omie=3037854759
        )
        self.assertEqual(conta_receber.cliente.codigo_cliente_omie, 101)
        self.assertEqual(conta_receber.conta_corrente, conta)
        self.assertEqual(conta_receber.categoria_principal, categoria)
        self.assertEqual(conta_receber.projeto, projeto)
        self.assertEqual(
            conta_receber.data_vencimento.isoformat(),
            "2020-10-20",
        )
        self.assertEqual(str(conta_receber.valor_documento), "500.00")
        self.assertEqual(str(conta_receber.valor_a_receber), "500.00")
        self.assertEqual(conta_receber.status_titulo, "CANCELADO")
        self.assertEqual(conta_receber.boleto["cGerado"], "N")
        self.assertEqual(conta_receber.codigo_pedido_omie, 3037850626)
        self.assertEqual(conta_receber.tipo_agrupamento, "I")
        lancamento = LancamentoContaCorrenteOmie.objects.get(
            codigo_lancamento_omie=4353703652
        )
        self.assertEqual(lancamento.conta_corrente, conta)
        self.assertEqual(lancamento.categoria_principal, categoria)
        self.assertEqual(lancamento.cliente_fornecedor.codigo_cliente_omie, 101)
        self.assertEqual(lancamento.projeto, projeto)
        self.assertEqual(lancamento.conta_receber, conta_receber)
        self.assertEqual(lancamento.data_lancamento.isoformat(), "2026-04-08")
        self.assertEqual(str(lancamento.valor_lancamento), "2289.50")
        self.assertEqual(lancamento.natureza, "R")
        self.assertEqual(lancamento.origem, "BAXR")
        self.assertEqual(lancamento.departamentos[0]["cCodDep"], "5476993662")
        self.assertEqual(categoria.dados_originais["descricao"], "Transferência")
